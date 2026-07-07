"""Shared tabular import/export helpers (CSV + XLSX).

Centralizes parsing of uploaded spreadsheets and generation of downloads so the
influencer, template, campaign and analytics endpoints share one implementation
instead of re-inlining openpyxl/csv logic.
"""
from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from datetime import date, datetime
from typing import Any

from fastapi import Response

from app.core.exceptions import BadRequestException

CSV_MEDIA_TYPE = "text/csv"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Cells openpyxl can write natively; everything else is coerced to str.
_NATIVE_CELL_TYPES = (str, int, float, bool, datetime, date, type(None))


def normalize_format(fmt: str | None) -> str:
    """Validate and normalize a requested export format ('csv' | 'xlsx')."""
    value = (fmt or "csv").strip().lower()
    if value not in ("csv", "xlsx"):
        raise BadRequestException("Unsupported export format; use 'csv' or 'xlsx'")
    return value


def parse_tabular(filename: str | None, content: bytes) -> list[dict[str, str]]:
    """Parse an uploaded .csv or .xlsx file into a list of lower-cased-key rows.

    Header names are trimmed and lower-cased; cell values are stringified and
    trimmed so downstream row mappers behave identically regardless of source
    format. Raises BadRequestException for unsupported extensions or unreadable
    files.
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx(content)
    if name.endswith(".csv"):
        return _parse_csv(content)
    raise BadRequestException("Only CSV or XLSX uploads are supported")


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    # utf-8-sig transparently strips a BOM written by Excel's "CSV UTF-8".
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict[str, str]] = []
    for raw in reader:
        rows.append({
            (k or "").strip().lower(): (v or "").strip()
            for k, v in raw.items()
            if k is not None
        })
    return rows


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a variety of types on bad files
        raise BadRequestException(f"Could not read XLSX file: {exc}") from exc

    ws = wb.active
    if ws is None:
        return []

    row_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        return []

    headers = [str(c).strip().lower() if c is not None else "" for c in header_row]
    rows: list[dict[str, str]] = []
    for values in row_iter:
        if values is None or all(v is None for v in values):
            continue  # skip fully blank rows
        row: dict[str, str] = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            v = values[i] if i < len(values) else None
            row[header] = str(v).strip() if v is not None else ""
        rows.append(row)
    wb.close()
    return rows


def _coerce_cell(value: Any) -> Any:
    if isinstance(value, _NATIVE_CELL_TYPES):
        return value
    # Enums, UUIDs, and anything exotic -> their string form.
    return str(getattr(value, "value", value))


def rows_to_csv_bytes(headers: list[str], rows: Iterable[Iterable[Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(["" if v is None else _stringify(v) for v in row])
    # utf-8-sig so Excel opens non-ASCII (Chinese) correctly on double-click.
    return buf.getvalue().encode("utf-8-sig")


def _stringify(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(getattr(value, "value", value))


def sheets_to_xlsx_bytes(
    sheets: list[tuple[str, list[str], Iterable[Iterable[Any]]]],
) -> bytes:
    """Build a multi-sheet workbook. Each sheet is (title, headers, rows)."""
    import openpyxl

    wb = openpyxl.Workbook()
    default = wb.active
    for idx, (title, headers, rows) in enumerate(sheets):
        ws = default if idx == 0 else wb.create_sheet()
        ws.title = (title or f"Sheet{idx + 1}")[:31]  # Excel caps titles at 31 chars
        ws.append(headers)
        for row in rows:
            ws.append([_coerce_cell(v) for v in row])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rows_to_xlsx_bytes(
    headers: list[str], rows: Iterable[Iterable[Any]], sheet_title: str = "Sheet1"
) -> bytes:
    return sheets_to_xlsx_bytes([(sheet_title, headers, rows)])


def tabular_response(
    *,
    fmt: str,
    filename_stem: str,
    headers: list[str],
    rows: list[list[Any]],
    sheet_title: str | None = None,
) -> Response:
    """Build a download Response in the requested format ('csv' | 'xlsx')."""
    fmt = normalize_format(fmt)
    if fmt == "xlsx":
        body = rows_to_xlsx_bytes(headers, rows, sheet_title or filename_stem)
        media_type = XLSX_MEDIA_TYPE
        filename = f"{filename_stem}.xlsx"
    else:
        body = rows_to_csv_bytes(headers, rows)
        media_type = CSV_MEDIA_TYPE
        filename = f"{filename_stem}.csv"
    return Response(
        content=body,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
