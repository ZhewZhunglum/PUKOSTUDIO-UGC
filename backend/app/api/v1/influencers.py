import csv
import io
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.exceptions import BadRequestException
from app.core.pagination import PaginationParams, get_pagination
from app.dependencies import get_current_user
from app.models.user import User
from app.schemas.email import EmailMessageResponse
from app.schemas.influencer import (
    BulkTagRequest,
    InfluencerCreate,
    InfluencerCRMActionRequest,
    InfluencerImportResponse,
    InfluencerListResponse,
    InfluencerResponse,
    InfluencerUpdate,
    TagCreate,
    TagResponse,
)
from app.services import influencer_service

router = APIRouter()


@router.get("", response_model=InfluencerListResponse)
async def list_influencers(
    search: str | None = Query(None),
    status: str | None = Query(None),
    niche: str | None = Query(None),
    platform: str | None = Query(None),
    source: str | None = Query(None),
    data_provider: str | None = Query(None),
    has_email: bool | None = Query(None),
    synced_after: datetime | None = Query(None),
    tag_id: uuid.UUID | None = Query(None),
    min_followers: int | None = Query(None),
    max_followers: int | None = Query(None),
    sort_by: str | None = Query(None),
    sort_order: str | None = Query("desc"),
    pagination: PaginationParams = Depends(get_pagination),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await influencer_service.list_influencers(
        db,
        team_id=current_user.team_id,
        params=pagination,
        search=search,
        status=status,
        niche=niche,
        platform=platform,
        source=source,
        data_provider=data_provider,
        has_email=has_email,
        synced_after=synced_after,
        tag_id=tag_id,
        min_followers=min_followers,
        max_followers=max_followers,
        sort_by=sort_by,
        sort_order=sort_order,
    )
    return result


@router.post("", response_model=InfluencerResponse, status_code=201)
async def create_influencer(
    data: InfluencerCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await influencer_service.create_influencer(db, current_user.team_id, data)


@router.post("/import", response_model=InfluencerImportResponse)
async def import_influencers(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise BadRequestException("Only CSV or XLSX uploads are supported")

    content = await file.read()

    if filename.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = [{k.strip().lower(): v for k, v in row.items()} for row in reader]

    result = await influencer_service.import_influencers_from_rows(
        db, current_user.team_id, rows
    )
    return result


@router.get("/export")
async def export_influencers(
    search: str | None = Query(None),
    status: str | None = Query(None),
    niche: str | None = Query(None),
    platform: str | None = Query(None),
    source: str | None = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.config import settings
    from app.core.pagination import PaginationParams

    BATCH_SIZE = 500

    async def generate_csv():
        header = io.StringIO()
        csv.writer(header).writerow([
            "name", "email", "niche", "country", "status",
            "platform", "username", "followers", "engagement_rate", "source", "created_at",
        ])
        yield header.getvalue()

        page = 1
        fetched = 0
        while fetched < settings.export_max_rows:
            result = await influencer_service.list_influencers(
                db,
                team_id=current_user.team_id,
                params=PaginationParams(page=page, per_page=BATCH_SIZE),
                search=search,
                status=status,
                niche=niche,
                platform=platform,
                source=source,
            )
            items = result["items"]
            if not items:
                break

            buf = io.StringIO()
            writer = csv.writer(buf)
            for inf in items:
                first_platform = inf.platforms[0] if inf.platforms else None
                writer.writerow([
                    inf.name,
                    inf.email or "",
                    inf.niche or "",
                    inf.country or "",
                    inf.status.value if inf.status else "",
                    first_platform.platform.value if first_platform else "",
                    first_platform.username if first_platform else "",
                    first_platform.followers if first_platform else "",
                    first_platform.engagement_rate if first_platform else "",
                    inf.source or "",
                    inf.created_at.isoformat() if inf.created_at else "",
                ])
            yield buf.getvalue()

            fetched += len(items)
            if len(items) < BATCH_SIZE:
                break
            page += 1

    return StreamingResponse(
        generate_csv(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=influencers.csv"},
    )


@router.get("/crm-summary")
async def get_influencer_crm_summary(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await influencer_service.get_crm_summary(db, current_user.team_id)


@router.post("/bulk-tag")
async def bulk_tag(
    data: BulkTagRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    count = await influencer_service.bulk_tag_influencers(
        db, current_user.team_id, data.influencer_ids, data.tag_ids
    )
    return {"updated": count}


# Tag management
@router.get("/tags/list", response_model=list[TagResponse])
async def list_tags(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select

    from app.models.influencer import Tag

    result = await db.execute(
        select(Tag).where(Tag.team_id == current_user.team_id).order_by(Tag.name)
    )
    return result.scalars().all()


@router.post("/tags", response_model=TagResponse, status_code=201)
async def create_tag(
    data: TagCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.models.influencer import Tag

    tag = Tag(team_id=current_user.team_id, name=data.name, color=data.color)
    db.add(tag)
    await db.flush()
    return tag


@router.get("/{influencer_id}/emails", response_model=list[EmailMessageResponse])
async def get_influencer_email_history(
    influencer_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await influencer_service.get_influencer_emails(
        db, influencer_id, current_user.team_id
    )


@router.post("/{influencer_id}/crm-action", response_model=InfluencerResponse)
async def apply_influencer_crm_action(
    influencer_id: uuid.UUID,
    data: InfluencerCRMActionRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await influencer_service.apply_crm_action(
        db,
        influencer_id,
        current_user.team_id,
        data.action,
        data.note,
    )


@router.post("/{influencer_id}/woto-refresh", response_model=InfluencerResponse)
async def refresh_influencer_from_woto(
    influencer_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await influencer_service.refresh_influencer_from_woto(
        db,
        influencer_id,
        current_user.team_id,
    )


@router.get("/{influencer_id}", response_model=InfluencerResponse)
async def get_influencer(
    influencer_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await influencer_service.get_influencer(db, influencer_id, current_user.team_id)


@router.put("/{influencer_id}", response_model=InfluencerResponse)
async def update_influencer(
    influencer_id: uuid.UUID,
    data: InfluencerUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await influencer_service.update_influencer(
        db, influencer_id, current_user.team_id, data
    )


@router.delete("/{influencer_id}", status_code=204)
async def delete_influencer(
    influencer_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await influencer_service.delete_influencer(db, influencer_id, current_user.team_id)
